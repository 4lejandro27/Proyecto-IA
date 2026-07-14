from openpyxl import load_workbook


def obtener_hojas_excel(ruta_archivo):

    libro = load_workbook(ruta_archivo)

    hojas = libro.sheetnames

    return hojas

def leer_hoja_excel(ruta_archivo, nombre_hoja):

    libro = load_workbook(ruta_archivo)

    hoja = libro[nombre_hoja]

    datos = []

    for fila in hoja.iter_rows(values_only=True):
        datos.append(list(fila))

    return datos